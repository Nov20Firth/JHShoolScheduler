# -*- coding: utf-8 -*-
"""國中排課的資料讀取、求解、驗證與 HTML／Excel 輸出函式。"""

import html
import json
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from pathlib import Path


# 開發資料夾可直接放置專案自己的套件，不必修改系統 Python。
LOCAL_PACKAGES = Path(__file__).with_name(".python-packages")
if LOCAL_PACKAGES.is_dir():
    sys.path.insert(0, str(LOCAL_PACKAGES))

from ortools.sat.python import cp_model
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from Parameters import (
    ALL_SLOTS,
    DAYS,
    DEFAULT_SCHOOL_INFO,
    EXAM_SUBJECTS,
    HOMEROOM_TEACHER_SUFFIX,
    MAX_EXAM_PERIODS_PER_DAY,
    MIN_EXAM_HOURS_FOR_GAP_RULE,
    normalize_subject,
    PERIODS,
    PERIOD_TIMES,
    PHYSICAL_EDUCATION_SUBJECT,
    REQUIRED_CLASS_HOURS,
    ROOM_BY_SUBJECT,
    SCHOOL_NAME_SUFFIX,
    SOLVER_MAX_TIME_SECONDS,
    SOLVER_NUM_SEARCH_WORKERS,
    WEIGHT_CROSS_SUBJECT,
    WEIGHT_EXAM_DAILY_OVERLOAD,
    WEIGHT_EXAM_PERIOD_5,
    WEIGHT_EXAM_PERIOD_7,
    WEIGHT_EXAM_TWO_DAY_GAP,
    WEIGHT_HOMEROOM_EARLY,
    WEIGHT_TEACHER_CONSECUTIVE_THREE,
    WEIGHT_TEACHER_LATE,
    WEIGHT_TEACHER_LUNCH,
)


if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")


@dataclass(frozen=True)
class Course:
    teacher: str
    subject: str
    hours: int
    klass: str
    role: str = "R"

    @property
    def key(self):
        return self.teacher, self.subject, self.klass, self.role


@dataclass
class BindingGroup:
    group_id: str
    mode: str
    day: str
    period: int | None
    special: tuple[str, str, str, str] | None = None

    @property
    def fixed_slot(self):
        return (self.day, self.period) if self.mode == "FIXED" else None


@dataclass
class BindingSet:
    set_id: str
    classes: list[str]
    subjects: list[tuple[str, int]]
    groups: list[BindingGroup] = field(default_factory=list)


def load_data_file(path):
    courses = []
    fixed_assignments = []
    binding_sets = {}
    teacher_unavailable = defaultdict(set)
    school_info = dict(DEFAULT_SCHOOL_INFO)

    with open(path, encoding="utf-8") as file:
        for raw_line in file:
            line = raw_line.strip()
            if not line:
                continue
            parts = line.split("|")
            tag = parts[0]

            if tag == "S":
                _, name, year, semester = parts
                school_info = {
                    "school_name": name,
                    "school_year": year,
                    "semester": semester,
                }
            elif tag == "C":
                _, teacher, subject, hours, klass = parts
                courses.append(Course(teacher, normalize_subject(subject), int(hours), klass))
            elif tag == "F":
                _, teacher, subject, klass, day, period = parts
                fixed_assignments.append((teacher, normalize_subject(subject), klass, day, int(period)))
            elif tag == "BK":
                _, set_id, class_text, subject_text = parts
                classes = class_text.split(",")
                subjects = [(normalize_subject(subject), int(hours))
                            for subject, hours in (item.split(":") for item in subject_text.split(","))]
                binding_sets[set_id] = BindingSet(set_id, classes, subjects)
            elif tag == "BT":
                _, set_id, group_id, mode, day, period, teacher, subject, klass = parts
                special = None if teacher == "-" else (teacher, normalize_subject(subject), klass, "S")
                binding_sets[set_id].groups.append(BindingGroup(
                    group_id,
                    mode,
                    day,
                    None if period == "-" else int(period),
                    special,
                ))
            elif tag == "U":
                _, teacher, day, period = parts
                teacher_unavailable[teacher].add((day, int(period)))

    special_counts = Counter(
        member
        for binding in binding_sets.values()
        for group in binding.groups
        if group.special
        for member in [group.special]
    )
    courses.extend(
        Course(teacher, subject, hours, klass, "S")
        for (teacher, subject, klass, _), hours in special_counts.items()
    )

    return {
        "courses": courses,
        "fixed_assignments": fixed_assignments,
        "binding_sets": list(binding_sets.values()),
        "teacher_unavailable": teacher_unavailable,
        "school_info": school_info,
    }


def print_data_summary(courses, binding_sets):
    regular_hours = defaultdict(int)
    special_hours = defaultdict(int)
    for course in courses:
        target = special_hours if course.role == "S" else regular_hours
        target[course.klass] += course.hours

    print("\n===== 資料摘要 =====")
    for klass in sorted(regular_hours):
        print(f"  {klass}班：{regular_hours[klass]} 堂")
    for klass in sorted(special_hours):
        print(f"  {klass}特生課表：{special_hours[klass]} 堂抽離課")
    groups = [group for binding in binding_sets for group in binding.groups]
    auto_count = sum(group.mode == "AUTO" for group in groups)
    fixed_count = len(groups) - auto_count
    print(f"  綁課群組：{len(groups)} 組（系統安排 {auto_count}、固定 {fixed_count}）\n")


def fixed_rows(binding_sets, fixed_assignments):
    rows = [(teacher, subject, klass, day, period, "F")
            for teacher, subject, klass, day, period in fixed_assignments]
    for binding in binding_sets:
        for group in binding.groups:
            if group.mode == "FIXED" and group.special:
                teacher, subject, klass, role = group.special
                rows.append((teacher, subject, klass, group.day, group.period, group.group_id))
    return rows


def run_diagnostics(courses, teacher_unavailable, binding_sets, fixed_assignments):
    issues = []
    duplicate_courses = Counter(
        (course.klass, course.subject) for course in courses if course.role == "R"
    )
    for (klass, subject), count in sorted(duplicate_courses.items()):
        if count > 1:
            issues.append(f"{klass}班的{subject}有 {count} 筆課程，每個班級＋科目只能有一筆。")
    regular_hours = defaultdict(int)
    for course in courses:
        if course.role == "R":
            regular_hours[course.klass] += course.hours
    for klass, hours in sorted(regular_hours.items()):
        if hours != REQUIRED_CLASS_HOURS:
            issues.append(
                f"{klass}班共有 {hours} 堂，不是 {REQUIRED_CLASS_HOURS} 堂。"
            )
    for teacher in sorted({course.teacher for course in courses}):
        needed = sum(course.hours for course in courses if course.teacher == teacher)
        available = len(ALL_SLOTS) - len(teacher_unavailable.get(teacher, set()))
        if needed > available:
            issues.append(f"{teacher} 需要 {needed} 堂，但只有 {available} 個可排時段。")

    rows = fixed_rows(binding_sets, fixed_assignments)
    slot_usage = defaultdict(list)
    for teacher, subject, klass, day, period, source in rows:
        slot_usage[day, period].append((teacher, klass, source))
        if (day, period) in teacher_unavailable.get(teacher, set()):
            issues.append(f"{teacher} 的固定課週{day}第{period}節落在不可排時段。")

    for (day, period), entries in slot_usage.items():
        teachers = [teacher for teacher, klass, source in entries]
        classes = [klass for teacher, klass, source in entries]
        for teacher, count in Counter(teachers).items():
            if count > 1:
                issues.append(f"週{day}第{period}節，{teacher} 有重複的固定課。")
        for klass, count in Counter(classes).items():
            if count > 1:
                issues.append(f"週{day}第{period}節，{klass}班有重複的固定課。")
    return issues


def build_and_solve(courses, teacher_unavailable, fixed_assignments, binding_sets):
    issues = run_diagnostics(courses, teacher_unavailable, binding_sets, fixed_assignments)
    if issues:
        print("\n排課前發現問題：")
        for issue in issues:
            print(f"  - {issue}")
        return None

    course_index = {course.key: index for index, course in enumerate(courses)}
    model = cp_model.CpModel()
    x = {}
    for index, course in enumerate(courses):
        for day, period in ALL_SLOTS:
            if (day, period) not in teacher_unavailable.get(course.teacher, set()):
                x[index, day, period] = model.new_bool_var(f"c{index}_{day}_{period}")

    def var(index, day, period):
        return x.get((index, day, period))

    for index, course in enumerate(courses):
        model.add(sum(var(index, day, period)
                      for day, period in ALL_SLOTS
                      if var(index, day, period) is not None) == course.hours)

    teacher_slots = defaultdict(list)
    class_slots = defaultdict(list)
    room_slots = defaultdict(list)
    for index, course in enumerate(courses):
        for day, period in ALL_SLOTS:
            value = var(index, day, period)
            if value is None:
                continue
            teacher_slots[course.teacher, day, period].append(value)
            class_slots[course.klass, day, period].append(value)
            room = ROOM_BY_SUBJECT.get(course.subject) if course.role == "R" else None
            if room:
                room_slots[room, day, period].append(value)

    for variables in teacher_slots.values():
        model.add(sum(variables) <= 1)
    for variables in class_slots.values():
        model.add(sum(variables) <= 1)
    for variables in room_slots.values():
        model.add(sum(variables) <= 1)

    # 每組綁課選一個時段；每個普通班在每組中選一科，再由模型同時決定交錯方式。
    course_by_class_subject = {
        (course.klass, course.subject): index
        for index, course in enumerate(courses)
        if course.role == "R"
    }
    group_choices = {}
    binding_assignments = {}
    binding_occurrences = defaultdict(list)
    for binding in binding_sets:
        regular_indices = {
            (klass, subject): course_by_class_subject[klass, subject]
            for klass in binding.classes
            for subject, hours in binding.subjects
        }
        for group in binding.groups:
            special_index = course_index[group.special] if group.special else None
            candidate_slots = [group.fixed_slot] if group.fixed_slot else ALL_SLOTS
            choices = []
            for day, period in candidate_slots:
                if special_index is not None and var(special_index, day, period) is None:
                    continue
                choice = model.new_bool_var(f"bind_{group.group_id}_{day}_{period}")
                group_choices[group.group_id, day, period] = choice
                choices.append(choice)
                if special_index is not None:
                    binding_occurrences[special_index, day, period].append(choice)
            if not choices:
                print(f"綁課群組 {group.group_id} 沒有共同可用時段。")
                return None
            model.add(sum(choices) == 1)

            for klass in binding.classes:
                assignments = []
                for subject, hours in binding.subjects:
                    index = regular_indices[klass, subject]
                    selected = model.new_bool_var(f"subject_{group.group_id}_{klass}_{subject}")
                    binding_assignments[group.group_id, klass, subject] = selected
                    assignments.append(selected)
                    for day, period in candidate_slots:
                        choice = group_choices.get((group.group_id, day, period))
                        if choice is None:
                            continue
                        course_var = var(index, day, period)
                        if course_var is None:
                            model.add(choice + selected <= 1)
                        else:
                            occurrence = model.new_bool_var(
                                f"occ_{group.group_id}_{klass}_{subject}_{day}_{period}"
                            )
                            model.add(occurrence <= choice)
                            model.add(occurrence <= selected)
                            model.add(occurrence >= choice + selected - 1)
                            binding_occurrences[index, day, period].append(occurrence)
                model.add(sum(assignments) == 1)

        for klass in binding.classes:
            for subject, hours in binding.subjects:
                model.add(sum(binding_assignments[group.group_id, klass, subject]
                              for group in binding.groups) == hours)

    for (index, day, period), occurrences in binding_occurrences.items():
        model.add(sum(occurrences) <= var(index, day, period))

    for teacher, subject, klass, day, period in fixed_assignments:
        index = course_index[teacher, subject, klass, "R"]
        value = var(index, day, period)
        if value is None:
            print(f"固定課無法套用：{teacher}／{subject}／{klass}／週{day}第{period}節")
            return None
        model.add(value == 1)

    # 同一門一般課同一天最多一堂；若使用者已在同一天固定多堂，
    # 當天上限才提高到明確固定的堂數，其餘日期仍維持一堂。
    fixed_daily_limits = Counter(
        (teacher, subject, klass, day)
        for teacher, subject, klass, day, period in fixed_assignments
    )
    for index, course in enumerate(courses):
        if course.role == "S":
            continue
        daily = {}
        for day in DAYS:
            day_vars = [var(index, day, period) for period in PERIODS
                        if var(index, day, period) is not None]
            daily[day] = sum(day_vars)
            daily_limit = max(
                1,
                fixed_daily_limits[course.teacher, course.subject, course.klass, day],
            )
            model.add(daily[day] <= daily_limit)
        if course.subject == PHYSICAL_EDUCATION_SUBJECT:
            for first, second in zip(DAYS, DAYS[1:]):
                model.add(daily[first] + daily[second] <= 1)

    penalties = []

    # 不同科綁課優先讓同一時段的班級選擇不同科目。
    for binding in binding_sets:
        if len(binding.subjects) < 2:
            continue
        for group in binding.groups:
            for subject, hours in binding.subjects:
                duplicate = model.new_int_var(0, len(binding.classes) - 1,
                                              f"cross_{group.group_id}_{subject}")
                model.add(duplicate >= sum(binding_assignments[group.group_id, klass, subject]
                                           for klass in binding.classes) - 1)
                penalties.append((duplicate, WEIGHT_CROSS_SUBJECT))

    # 每週至少三堂的考科，盡量不要連續兩天都沒課。
    for index, course in enumerate(courses):
        if (course.role == "S" or course.subject not in EXAM_SUBJECTS
                or course.hours < MIN_EXAM_HOURS_FOR_GAP_RULE):
            continue
        used = {}
        for day in DAYS:
            day_vars = [var(index, day, period) for period in PERIODS
                        if var(index, day, period) is not None]
            if day_vars:
                used[day] = model.new_bool_var(f"used_{index}_{day}")
                model.add_max_equality(used[day], day_vars)
            else:
                used[day] = model.new_constant(0)
        for first, second in zip(DAYS, DAYS[1:]):
            gap = model.new_bool_var(f"gap_{index}_{first}_{second}")
            model.add(used[first] + used[second] + gap >= 1)
            penalties.append((gap, WEIGHT_EXAM_TWO_DAY_GAP))

    # 每班每天考科超過五堂時才扣分。
    regular_classes = sorted({course.klass for course in courses if course.role == "R"})
    for klass in regular_classes:
        for day in DAYS:
            exam_vars = []
            for index, course in enumerate(courses):
                if course.role == "R" and course.klass == klass and course.subject in EXAM_SUBJECTS:
                    exam_vars.extend(var(index, day, period) for period in PERIODS
                                     if var(index, day, period) is not None)
            overload = model.new_int_var(0, 2, f"exam_{klass}_{day}")
            model.add(overload >= sum(exam_vars) - MAX_EXAM_PERIODS_PER_DAY)
            penalties.append((overload, WEIGHT_EXAM_DAILY_OVERLOAD))

    # 考科盡量避開下午第 5、7 節；第 7 節的優先度較高。
    for index, course in enumerate(courses):
        if course.role == "S" or course.subject not in EXAM_SUBJECTS:
            continue
        for day in DAYS:
            for period, weight in [
                (5, WEIGHT_EXAM_PERIOD_5),
                (7, WEIGHT_EXAM_PERIOD_7),
            ]:
                value = var(index, day, period)
                if value is not None:
                    penalties.append((value, weight))

    # 教師條件依權重排序：跨午休、連續三堂、導師連上一二節、連上六七節。
    for teacher in sorted({course.teacher for course in courses}):
        for day in DAYS:
            for periods, weight, label in [
                ((4, 5), WEIGHT_TEACHER_LUNCH, "lunch"),
                ((1, 2, 3), WEIGHT_TEACHER_CONSECUTIVE_THREE, "triple1"),
                ((2, 3, 4), WEIGHT_TEACHER_CONSECUTIVE_THREE, "triple2"),
                ((3, 4, 5), WEIGHT_TEACHER_CONSECUTIVE_THREE, "triple3"),
                ((4, 5, 6), WEIGHT_TEACHER_CONSECUTIVE_THREE, "triple4"),
                ((5, 6, 7), WEIGHT_TEACHER_CONSECUTIVE_THREE, "triple5"),
                ((6, 7), WEIGHT_TEACHER_LATE, "late"),
            ]:
                slot_vars = [teacher_slots.get((teacher, day, period), []) for period in periods]
                if any(not values for values in slot_vars):
                    continue
                violation = model.new_bool_var(f"{label}_{teacher}_{day}")
                model.add(sum(sum(values) for values in slot_vars) - (len(periods) - 1) <= violation)
                penalties.append((violation, weight))
            if teacher.endswith(HOMEROOM_TEACHER_SUFFIX):
                first = teacher_slots.get((teacher, day, 1), [])
                second = teacher_slots.get((teacher, day, 2), [])
                if first and second:
                    violation = model.new_bool_var(f"homeroom_early_{teacher}_{day}")
                    model.add(sum(first) + sum(second) - 1 <= violation)
                    penalties.append((violation, WEIGHT_HOMEROOM_EARLY))

    model.minimize(sum(variable * weight for variable, weight in penalties))

    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = SOLVER_MAX_TIME_SECONDS
    solver.parameters.num_search_workers = SOLVER_NUM_SEARCH_WORKERS
    status = solver.solve(model)
    if status == cp_model.INFEASIBLE:
        print("\n找不到符合所有硬性條件的課表。")
        return None
    if status == cp_model.UNKNOWN:
        print(
            f"\n求解已達 {SOLVER_MAX_TIME_SECONDS} 秒上限，目前尚未找到課表；"
            "這不代表硬性條件一定無解。"
        )
        return None
    if status == cp_model.MODEL_INVALID:
        print("\n排課模型無效，請檢查輸入資料或程式限制。")
        return None

    result = []
    for index, course in enumerate(courses):
        for day, period in ALL_SLOTS:
            value = var(index, day, period)
            if value is not None and solver.value(value):
                result.append((course.teacher, course.subject, course.klass, day, period, course.role))

    binding_schedule = {}
    for (group_id, day, period), choice in group_choices.items():
        if solver.value(choice):
            assigned = [
                (klass, subject)
                for (candidate_group, klass, subject), selected in binding_assignments.items()
                if candidate_group == group_id and solver.value(selected)
            ]
            binding_schedule[group_id] = (day, period, assigned)

    print(f"求解完成：{solver.status_name(status)}，目標值 {solver.objective_value:g}")
    return result, binding_schedule


def group_result_by_entity(result, entity_index, cell_text):
    tables = {}
    for entity in sorted({row[entity_index] for row in result}):
        table = {period: {day: ("", "") for day in DAYS} for period in PERIODS}
        for teacher, subject, klass, day, period, role in result:
            row = (teacher, subject, klass, day, period, role)
            if row[entity_index] == entity:
                table[period][day] = cell_text(teacher, subject, klass)
        tables[entity] = table
    return tables


def teacher_sheet_label(name):
    return name if name.endswith(("老師", "導師", "教師")) else f"{name}\n老師"


class TimetableStyle:
    FONT = "Microsoft JhengHei"
    LINE = "777777"
    STRONG = "000000"
    HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
    SIDE_FILL = PatternFill("solid", fgColor="F2F2F2")
    LABEL_FILL = PatternFill("solid", fgColor="EEEEEE")
    CHANGED_FILL = PatternFill("solid", fgColor="C6EFCE")
    CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    THIN_SIDE = Side(style="thin", color="777777")
    STRONG_SIDE = Side(style="medium", color="000000")
    THIN = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
    HEADER = Border(
        left=STRONG_SIDE,
        right=STRONG_SIDE,
        top=STRONG_SIDE,
        bottom=STRONG_SIDE,
    )


def split_subject_name(subject):
    return subject if len(subject) <= 2 else f"{subject[:2]}\n{subject[2:]}"


def format_school_name(name):
    suffix = SCHOOL_NAME_SUFFIX
    return f"{name[:-len(suffix)]}\n{suffix}" if name.endswith(suffix) and name != suffix else name


def style_side_panel(ws, column):
    medium = Side(style="medium", color=TimetableStyle.STRONG)
    for row in range(1, 16):
        cell = ws.cell(row, column)
        cell.fill = TimetableStyle.SIDE_FILL
        cell.border = Border(
            left=medium,
            right=medium,
            top=medium if row == 1 else Side(),
            bottom=medium if row == 15 else Side(),
        )


def write_timetable_sheet(ws, school_info, kind, entity_label, table, highlighted_slots=None):
    style = TimetableStyle
    highlighted_slots = highlighted_slots or set()
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:A15")
    ws["A1"] = (f"{format_school_name(school_info['school_name'])}\n"
                f"{school_info['school_year']}學年度\n第{school_info['semester']}學期\n{kind}")
    ws["A1"].font = Font(name=style.FONT, bold=True, size=10)
    ws["A1"].alignment = style.CENTER
    style_side_panel(ws, 1)

    ws.merge_cells("I1:I15")
    ws["I1"] = entity_label
    ws["I1"].font = Font(name=style.FONT, bold=True, size=12)
    ws["I1"].alignment = style.CENTER
    style_side_panel(ws, 9)

    for column, value in enumerate(["節次", "時間", *DAYS], start=2):
        cell = ws.cell(1, column, value)
        cell.fill = style.HEADER_FILL
        cell.font = Font(name=style.FONT, bold=True, size=12 if column >= 4 else 10)
        cell.alignment = style.CENTER
        cell.border = style.HEADER

    for index, period in enumerate(PERIODS):
        subject_row = 2 + index * 2
        detail_row = subject_row + 1
        ws.merge_cells(start_row=subject_row, start_column=2, end_row=detail_row, end_column=2)
        ws.merge_cells(start_row=subject_row, start_column=3, end_row=detail_row, end_column=3)

        ws.cell(subject_row, 2, period)
        start, end = PERIOD_TIMES[period]
        ws.cell(subject_row, 3, f"{start}\n-\n{end}")
        for column in (2, 3):
            cell = ws.cell(subject_row, column)
            cell.fill = style.LABEL_FILL
            cell.font = Font(name=style.FONT, bold=column == 2, size=12 if column == 2 else 10)
            cell.alignment = style.CENTER

        for column, day in enumerate(DAYS, start=4):
            subject, detail = table[period][day]
            subject_cell = ws.cell(subject_row, column, split_subject_name(subject))
            detail_cell = ws.cell(detail_row, column, detail)
            if (day, period) in highlighted_slots:
                subject_cell.fill = detail_cell.fill = style.CHANGED_FILL
            subject_cell.font = Font(name=style.FONT, bold=True, size=11)
            detail_cell.font = Font(name=style.FONT, size=9, color="333333")
            subject_cell.alignment = detail_cell.alignment = style.CENTER

        for row in (subject_row, detail_row):
            for column in range(2, 9):
                ws.cell(row, column).border = style.THIN
        ws.row_dimensions[subject_row].height = 52
        ws.row_dimensions[detail_row].height = 28

    ws.row_dimensions[1].height = 28
    for column, width in {"A": 11.5, "B": 5, "C": 8, "D": 8.2, "E": 8.2,
                          "F": 8.2, "G": 8.2, "H": 8.2, "I": 9}.items():
        ws.column_dimensions[column].width = width
    ws.print_area = "A1:I15"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.blackAndWhite = not bool(highlighted_slots)
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True


def export_to_excel(
        result,
        school_info,
        binding_schedule,
        filename,
        highlighted_class_slots=None,
        highlighted_teacher_slots=None,
):
    highlighted_class_slots = highlighted_class_slots or {}
    highlighted_teacher_slots = highlighted_teacher_slots or {}
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    regular = [row for row in result if row[5] == "R"]
    special = [row for row in result if row[5] == "S"]

    class_tables = group_result_by_entity(
        regular, 2, lambda teacher, subject, klass: (subject, teacher)
    )
    for klass, table in class_tables.items():
        sheet = workbook.create_sheet(f"{klass}班"[:31])
        write_timetable_sheet(
            sheet,
            school_info,
            "班級課程表",
            f"{klass}班",
            table,
            highlighted_class_slots.get(klass),
        )

    special_tables = group_result_by_entity(
        special, 2, lambda teacher, subject, klass: (subject, teacher)
    )
    for klass, table in special_tables.items():
        sheet = workbook.create_sheet(f"{klass}班"[:31])
        write_timetable_sheet(sheet, school_info, "特生課程表", f"{klass}班", table)

    teacher_tables = group_result_by_entity(
        result, 0, lambda teacher, subject, klass: (subject, f"{klass}班")
    )
    for teacher, table in teacher_tables.items():
        sheet = workbook.create_sheet(teacher[:31])
        write_timetable_sheet(
            sheet,
            school_info,
            "教師課程表",
            teacher_sheet_label(teacher),
            table,
            highlighted_teacher_slots.get(teacher),
        )

    binding_sheet = workbook.create_sheet("綁課群組")
    binding_sheet.append(["群組", "星期", "節次", "普通班科目"])
    for group_id in sorted(binding_schedule):
        day, period, assigned = binding_schedule[group_id]
        content = "、".join(f"{klass} {subject}" for klass, subject in assigned)
        binding_sheet.append([group_id, day, period, content])
    binding_sheet.freeze_panes = "A2"
    binding_sheet.column_dimensions["A"].width = 12
    binding_sheet.column_dimensions["B"].width = 10
    binding_sheet.column_dimensions["C"].width = 10
    binding_sheet.column_dimensions["D"].width = 42

    try:
        workbook.save(filename)
    except PermissionError:
        print(f"\n無法寫入 {filename}：檔案可能正在 Excel 中開啟，請關閉後再試。")
        return None
    return filename


ADJUSTMENT_HTML_TEMPLATE = r'''<!doctype html>
<html lang="zh-Hant">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>__TITLE__</title>
<style>
:root {
  color-scheme: light;
  --ink: #243047;
  --muted: #64748b;
  --line: #cbd5e1;
  --panel: #f8fafc;
  --brand: #2563eb;
  --changed: #d9f2e6;
  --locked: #e5e7eb;
  font-family: "Microsoft JhengHei", "Noto Sans TC", sans-serif;
}
* { box-sizing: border-box; }
body { margin: 0; color: var(--ink); background: #eef2f7; }
header { background: #fff; border-bottom: 1px solid var(--line); padding: 20px 24px; }
h1 { margin: 0 0 6px; font-size: 24px; }
header p { margin: 0; color: var(--muted); }
main { max-width: 1280px; margin: 0 auto; padding: 20px; }
.instructions, .toolbar, .problems { background: #fff; border: 1px solid var(--line); border-radius: 12px; }
.instructions { padding: 14px 18px; margin-bottom: 14px; line-height: 1.75; }
.instructions strong { color: #b42318; }
.toolbar { display: flex; flex-wrap: wrap; gap: 10px; align-items: center; padding: 12px; margin-bottom: 14px; }
button { border: 1px solid var(--line); border-radius: 8px; background: #fff; color: var(--ink); padding: 9px 14px; font: inherit; cursor: pointer; }
button:hover { background: #f1f5f9; }
button.primary { background: var(--brand); border-color: var(--brand); color: #fff; }
button.primary:hover { background: #1d4ed8; }
button:disabled { opacity: .45; cursor: not-allowed; }
.counter { margin-left: auto; font-weight: 700; }
.tabs { display: flex; flex-wrap: wrap; gap: 8px; margin: 0 0 12px; }
.tab.active { background: var(--brand); border-color: var(--brand); color: #fff; }
.view-toggle.active { background: #0f172a; border-color: #0f172a; color: #fff; }
.table-wrap { overflow-x: auto; background: #fff; border: 1px solid var(--line); border-radius: 12px; padding: 12px; }
table { border-collapse: separate; border-spacing: 0; width: 100%; min-width: 900px; table-layout: fixed; }
th, td { border-right: 1px solid var(--line); border-bottom: 1px solid var(--line); padding: 6px; text-align: center; }
tr:first-child th { border-top: 1px solid var(--line); }
th:first-child, td:first-child { border-left: 1px solid var(--line); }
th { background: #e2e8f0; }
th:first-child { width: 72px; }
.slot { height: 92px; background: #fff; vertical-align: middle; }
.course { min-height: 74px; border: 2px solid #94a3b8; border-radius: 9px; background: #fff; padding: 8px 5px; display: flex; flex-direction: column; justify-content: center; gap: 5px; user-select: none; }
.course[draggable="true"] { cursor: grab; }
.course[draggable="true"]:active { cursor: grabbing; }
.course.changed { background: var(--changed); border-color: #4f9f78; }
.course.locked { background: var(--locked); border-color: #9ca3af; cursor: not-allowed; }
.course.binding-movable { background: #eff6ff; border-color: #3b82f6; }
.course.dragging { opacity: .35; }
.subject { font-size: 17px; font-weight: 800; }
.teacher { font-size: 13px; color: #475569; }
.lock { font-size: 11px; color: #6b7280; }
.slot.drop-target .course { outline: 4px solid #93c5fd; outline-offset: 1px; }
.problems { margin: 0 0 14px; padding: 14px 18px; }
.problems.ok { border-color: #86efac; background: #f0fdf4; }
.problems.bad { position: sticky; top: 8px; z-index: 20; border-color: #ef4444; background: #fff1f2; box-shadow: 0 8px 24px rgba(127, 29, 29, .22); }
.problems h2 { margin: 0 0 8px; font-size: 17px; }
.problems ul { margin: 0; padding-left: 24px; }
.legend { display: inline-flex; align-items: center; gap: 7px; color: var(--muted); font-size: 13px; }
.swatch { width: 18px; height: 18px; border: 1px solid #94a3b8; border-radius: 4px; display: inline-block; }
.swatch.changed { background: var(--changed); }
.swatch.locked { background: var(--locked); }
.swatch.binding-movable { background: #eff6ff; border-color: #3b82f6; }
@media (max-width: 700px) {
  header, main { padding: 14px; }
  .counter { width: 100%; margin-left: 0; }
}
</style>
</head>
<body>
<header>
  <h1 id="page-title">暫定課表微調</h1>
  <p>在班級課表中拖曳交換課程，頁面會即時檢查衝突；完成後請儲存微調資料。</p>
</header>
<main>
  <section class="instructions">
    <b>操作方式：</b>先選擇班級，再把一張「科目＋教師」課程卡拖到同班另一張卡片上，兩堂課會交換時段。
    藍框卡片是系統安排的綁課，可拖到同一天、同一綁課組的不同科綁課上；其他班級與特生課會整組同步交換。
    灰色卡片是固定課或固定綁課，不能移動。淡綠色代表已更動。
    「教師課表」可供查閱及確認衝堂，但不提供拖曳。
    若要跨日交換、變更綁課科目組合或參與班級，請回到資料整理頁修改綁課設定。
    <strong>完成後務必按「儲存微調資料」</strong>，瀏覽器會下載 <code>3b_data.json</code>，請妥善保存。
  </section>
  <section class="toolbar">
    <button id="class-view" class="view-toggle active" type="button">班級課表</button>
    <button id="teacher-view" class="view-toggle" type="button">教師課表（唯讀）</button>
    <button id="undo" type="button">復原上一步</button>
    <button id="reset" type="button">全部還原</button>
    <button id="save" class="primary" type="button">儲存微調資料</button>
    <span class="legend"><span class="swatch changed"></span>已更動</span>
    <span class="legend"><span class="swatch binding-movable"></span>系統綁課（整組可換）</span>
    <span class="legend"><span class="swatch locked"></span>固定／固定綁課</span>
    <span class="counter" id="counter">已更動 0 堂</span>
  </section>
  <section class="problems ok" id="problems" aria-live="assertive"></section>
  <nav class="tabs" id="tabs" aria-label="班級"></nav>
  <section class="table-wrap"><table id="schedule"></table></section>
</main>
<script id="schedule-data" type="application/json">__DATA__</script>
<script>
(() => {
  "use strict";
  const source = JSON.parse(document.getElementById("schedule-data").textContent);
  const original = source.result.map(item => ({...item}));
  let current = source.result.map(item => ({...item}));
  let history = [];
  let viewMode = "class";
  let activeClass = [...new Set(current.filter(x => x.role === "R").map(x => x.klass))].sort()[0];
  let activeTeacher = [...new Set(current.map(x => x.teacher))].sort()[0];
  const originalById = new Map(original.map(item => [item.id, item]));
  const dayOrder = new Map(source.days.map((day, index) => [day, index]));

  const changed = item => {
    const before = originalById.get(item.id);
    return before.day !== item.day || before.period !== item.period;
  };
  const changedCount = () => current.filter(changed).length;
  const isAutoBinding = item => Boolean(
    item.binding_mode === "AUTO" && item.binding_set_id && item.binding_group_id
  );
  const isMovable = item => !item.locked_reason || isAutoBinding(item);

  function renderTabs() {
    const tabs = document.getElementById("tabs");
    tabs.replaceChildren();
    const entities = viewMode === "class"
      ? [...new Set(current.filter(x => x.role === "R").map(x => x.klass))].sort()
      : [...new Set(current.map(x => x.teacher))].sort();
    const activeEntity = viewMode === "class" ? activeClass : activeTeacher;
    for (const entity of entities) {
      const button = document.createElement("button");
      button.type = "button";
      button.className = "tab" + (entity === activeEntity ? " active" : "");
      button.textContent = viewMode === "class" ? `${entity}班` : entity;
      button.addEventListener("click", () => {
        if (viewMode === "class") activeClass = entity;
        else activeTeacher = entity;
        render();
      });
      tabs.append(button);
    }
  }

  function cardFor(item, editable) {
    const card = document.createElement("div");
    card.className = "course";
    if (changed(item)) card.classList.add("changed");
    if (isAutoBinding(item)) card.classList.add("binding-movable");
    else if (item.locked_reason) card.classList.add("locked");
    card.draggable = editable && isMovable(item);
    card.dataset.id = item.id;

    const subject = document.createElement("div");
    subject.className = "subject";
    subject.textContent = item.subject;
    const teacher = document.createElement("div");
    teacher.className = "teacher";
    teacher.textContent = viewMode === "class" ? item.teacher : `${item.klass}班`;
    card.append(subject, teacher);
    if (isAutoBinding(item)) {
      const lock = document.createElement("div");
      lock.className = "lock";
      lock.textContent = "↔ 系統綁課（同日整組互換）";
      card.append(lock);
    } else if (item.locked_reason) {
      const lock = document.createElement("div");
      lock.className = "lock";
      lock.textContent = `🔒 ${item.locked_reason}`;
      card.append(lock);
    }

    if (editable) {
      card.addEventListener("dragstart", event => {
        if (!isMovable(item)) { event.preventDefault(); return; }
        event.dataTransfer.effectAllowed = "move";
        event.dataTransfer.setData("text/plain", item.id);
        requestAnimationFrame(() => card.classList.add("dragging"));
      });
      card.addEventListener("dragend", () => card.classList.remove("dragging"));
    }
    return card;
  }

  function renderSchedule() {
    const table = document.getElementById("schedule");
    table.replaceChildren();
    const header = document.createElement("tr");
    const corner = document.createElement("th");
    corner.textContent = "節次";
    header.append(corner);
    for (const day of source.days) {
      const th = document.createElement("th");
      th.textContent = `星期${day}`;
      header.append(th);
    }
    table.append(header);

    for (const period of source.periods) {
      const row = document.createElement("tr");
      const label = document.createElement("th");
      const times = source.period_times[String(period)] || [];
      label.textContent = times.length === 2 ? `第${period}節\n${times[0]}–${times[1]}` : `第${period}節`;
      label.style.whiteSpace = "pre-line";
      row.append(label);
      for (const day of source.days) {
        const cell = document.createElement("td");
        cell.className = "slot";
        const item = current.find(x =>
          x.day === day && x.period === period &&
          (viewMode === "class"
            ? x.role === "R" && x.klass === activeClass
            : x.teacher === activeTeacher)
        );
        if (item) {
          const editable = viewMode === "class";
          cell.append(cardFor(item, editable));
          if (editable) {
            cell.addEventListener("dragover", event => {
              if (isMovable(item)) { event.preventDefault(); cell.classList.add("drop-target"); }
            });
            cell.addEventListener("dragleave", () => cell.classList.remove("drop-target"));
            cell.addEventListener("drop", event => {
              event.preventDefault();
              cell.classList.remove("drop-target");
              swapCourses(event.dataTransfer.getData("text/plain"), item.id);
            });
          }
        }
        row.append(cell);
      }
      table.append(row);
    }
  }

  function swapCourses(sourceId, targetId) {
    if (!sourceId || sourceId === targetId) return;
    const first = current.find(x => x.id === sourceId);
    const second = current.find(x => x.id === targetId);
    if (!first || !second || first.klass !== second.klass || first.role !== "R" || second.role !== "R") {
      alert("目前微調只支援同一班內交換課程。");
      return;
    }
    const firstAuto = isAutoBinding(first);
    const secondAuto = isAutoBinding(second);
    if (firstAuto || secondAuto) {
      if (!firstAuto || !secondAuto) {
        alert("系統綁課只能與同一綁課組內的另一堂系統綁課互換。");
        return;
      }
      if (first.binding_set_id !== second.binding_set_id) {
        alert("兩堂課不屬於同一綁課組，不能互換。");
        return;
      }
      if (first.binding_group_id === second.binding_group_id) return;
      if (first.day !== second.day) {
        alert("系統綁課目前只開放同一天內互換。");
        return;
      }
      if (first.subject === second.subject) {
        alert("請選擇同一綁課組內的不同科目互換。");
        return;
      }
      const firstMembers = current.filter(item => item.binding_group_id === first.binding_group_id);
      const secondMembers = current.filter(item => item.binding_group_id === second.binding_group_id);
      if (!firstMembers.length || !secondMembers.length) {
        alert("找不到完整綁課資料，請重新產生暫定課表。");
        return;
      }
      const firstSlot = [first.day, first.period];
      const secondSlot = [second.day, second.period];
      history.push(current.map(item => ({...item})));
      for (const item of firstMembers) [item.day, item.period] = secondSlot;
      for (const item of secondMembers) [item.day, item.period] = firstSlot;
      render();
      return;
    }
    if (first.locked_reason || second.locked_reason) {
      alert("固定課與固定綁課不能移動。");
      return;
    }
    history.push(current.map(item => ({...item})));
    [first.day, second.day] = [second.day, first.day];
    [first.period, second.period] = [second.period, first.period];
    render();
  }

  function validate() {
    const problems = [];
    const teacherSlots = new Map();
    const roomSlots = new Map();
    const classCourseDays = new Map();
    const fixedDailyLimits = new Map();
    for (const [teacher, subject, klass, day] of source.fixed_assignments || []) {
      const key = `${klass}|${teacher}|${subject}|${day}`;
      fixedDailyLimits.set(key, (fixedDailyLimits.get(key) || 0) + 1);
    }
    const unavailable = new Map(Object.entries(source.teacher_unavailable)
      .map(([teacher, slots]) => [teacher, new Set(slots.map(slot => `${slot[0]}|${slot[1]}`))]));

    for (const item of current) {
      const teacherKey = `${item.teacher}|${item.day}|${item.period}`;
      if (teacherSlots.has(teacherKey)) {
        const other = teacherSlots.get(teacherKey);
        problems.push(`教師衝堂：${item.teacher} 在週${item.day}第${item.period}節同時教授 ${other.klass}班與${item.klass}班。`);
      } else teacherSlots.set(teacherKey, item);

      if (unavailable.get(item.teacher)?.has(`${item.day}|${item.period}`))
        problems.push(`不可排時段：${item.teacher} 不能在週${item.day}第${item.period}節上課。`);

      if (item.role === "R" && source.room_by_subject[item.subject]) {
        const room = source.room_by_subject[item.subject];
        const roomKey = `${room}|${item.day}|${item.period}`;
        if (roomSlots.has(roomKey)) problems.push(`專科教室衝突：${room} 在週${item.day}第${item.period}節重複使用。`);
        else roomSlots.set(roomKey, item);
      }

      if (item.role === "R") {
        const dayKey = `${item.klass}|${item.teacher}|${item.subject}|${item.day}`;
        classCourseDays.set(dayKey, (classCourseDays.get(dayKey) || 0) + 1);
      }
    }

    for (const [dayKey, count] of classCourseDays) {
      const [klass, teacher, subject, day] = dayKey.split("|");
      const limit = Math.max(1, fixedDailyLimits.get(dayKey) || 0);
      if (count > limit)
        problems.push(`同科同日重複：${klass}班的${subject}在週${day}排了 ${count} 堂，允許上限為 ${limit} 堂。`);
    }

    const peByClass = new Map();
    for (const item of current.filter(x => x.role === "R" && x.subject === source.physical_education_subject)) {
      if (!peByClass.has(item.klass)) peByClass.set(item.klass, new Set());
      peByClass.get(item.klass).add(dayOrder.get(item.day));
    }
    for (const [klass, days] of peByClass) {
      for (const index of days) if (days.has(index + 1)) {
        problems.push(`體育連續兩天：${klass}班的體育排在相鄰上課日。`);
        break;
      }
    }

    const bindingSlots = new Map();
    for (const item of current.filter(x => x.binding_group_id)) {
      if (!bindingSlots.has(item.binding_group_id)) bindingSlots.set(item.binding_group_id, new Set());
      bindingSlots.get(item.binding_group_id).add(`${item.day}|${item.period}`);
    }
    for (const [groupId, slots] of bindingSlots) {
      if (slots.size !== 1) problems.push(`綁課群組 ${groupId} 未完整同步移動。`);
    }
    const originalSlotsBySetDay = new Map();
    const currentSlotsBySetDay = new Map();
    for (const [groupId, info] of Object.entries(source.binding_schedule)) {
      const key = `${info.set_id}|${info.day}`;
      if (!originalSlotsBySetDay.has(key)) originalSlotsBySetDay.set(key, []);
      originalSlotsBySetDay.get(key).push(info.period);
      const member = current.find(item => item.binding_group_id === groupId);
      if (!member) {
        problems.push(`綁課群組 ${groupId} 缺少課程資料。`);
        continue;
      }
      if (info.mode === "FIXED" && (member.day !== info.day || member.period !== info.period))
        problems.push(`固定綁課群組 ${groupId} 不能移動。`);
      if (info.mode === "AUTO" && member.day !== info.day)
        problems.push(`系統綁課群組 ${groupId} 只能在原星期內互換。`);
      const currentKey = `${info.set_id}|${member.day}`;
      if (!currentSlotsBySetDay.has(currentKey)) currentSlotsBySetDay.set(currentKey, []);
      currentSlotsBySetDay.get(currentKey).push(member.period);
    }
    for (const [key, periods] of originalSlotsBySetDay) {
      const before = [...periods].sort((a, b) => a - b).join(",");
      const after = [...(currentSlotsBySetDay.get(key) || [])].sort((a, b) => a - b).join(",");
      if (before !== after) problems.push(`綁課組 ${key.split("|")[0]} 的原有時段被改變。`);
    }
    return [...new Set(problems)];
  }

  function currentBindingSchedule() {
    return Object.fromEntries(Object.entries(source.binding_schedule).map(([groupId, info]) => {
      const member = current.find(item => item.binding_group_id === groupId);
      return [groupId, {...info, day: member?.day || info.day, period: member?.period || info.period}];
    }));
  }

  function renderProblems(problems) {
    const box = document.getElementById("problems");
    box.replaceChildren();
    box.className = "problems " + (problems.length ? "bad" : "ok");
    const title = document.createElement("h2");
    title.textContent = problems.length ? `目前發現 ${problems.length} 個衝突` : "目前未發現明顯硬限制衝突";
    box.append(title);
    if (problems.length) {
      const list = document.createElement("ul");
      for (const problem of problems) { const li = document.createElement("li"); li.textContent = problem; list.append(li); }
      box.append(list);
    } else {
      const note = document.createElement("div");
      note.textContent = "請再查看班級與教師課表，確認調整結果符合實際需求。";
      box.append(note);
    }
  }

  function render() {
    document.getElementById("class-view").classList.toggle("active", viewMode === "class");
    document.getElementById("teacher-view").classList.toggle("active", viewMode === "teacher");
    renderTabs();
    renderSchedule();
    renderProblems(validate());
    document.getElementById("counter").textContent = `已更動 ${changedCount()} 堂`;
    document.getElementById("undo").disabled = history.length === 0;
  }

  document.getElementById("class-view").addEventListener("click", () => {
    viewMode = "class";
    render();
  });
  document.getElementById("teacher-view").addEventListener("click", () => {
    viewMode = "teacher";
    render();
  });

  document.getElementById("undo").addEventListener("click", () => {
    if (!history.length) return;
    current = history.pop();
    render();
  });
  document.getElementById("reset").addEventListener("click", () => {
    if (!changedCount()) return;
    if (!confirm("確定要放棄全部微調嗎？")) return;
    history.push(current.map(item => ({...item})));
    current = original.map(item => ({...item}));
    render();
  });
  document.getElementById("save").addEventListener("click", () => {
    const problems = validate();
    if (problems.length) { alert("目前仍有硬限制衝突，請先修正紅色警告。") ; return; }
    if (!changedCount()) { alert("目前尚未更動任何課程。"); return; }
    const payload = {
      format: "junior-high-timetable-adjustment-v3",
      created_at: new Date().toISOString(),
      school_info: source.school_info,
      days: source.days,
      periods: source.periods,
      period_times: source.period_times,
      teacher_unavailable: source.teacher_unavailable,
      room_by_subject: source.room_by_subject,
      physical_education_subject: source.physical_education_subject,
      fixed_assignments: source.fixed_assignments || original
        .filter(item => item.locked_reason === "固定課")
        .map(item => [item.teacher, item.subject, item.klass, item.day, item.period]),
      original_result: original,
      adjusted_result: current,
      binding_schedule: currentBindingSchedule()
    };
    const blob = new Blob([JSON.stringify(payload, null, 2)], {type: "application/json;charset=utf-8"});
    const link = document.createElement("a");
    link.href = URL.createObjectURL(blob);
    link.download = "3b_data.json";
    document.body.append(link);
    link.click();
    link.remove();
    setTimeout(() => URL.revokeObjectURL(link.href), 1000);
    alert("已下載 3b_data.json，請妥善保存這份微調資料。");
  });

  document.getElementById("page-title").textContent = `${source.school_info.school_name}－暫定課表微調`;
  render();
})();
</script>
</body>
</html>
'''


def export_to_adjustment_html(
        result,
        school_info,
        binding_schedule,
        binding_sets,
        fixed_assignments,
        teacher_unavailable,
        filename,
):
    """輸出可在瀏覽器中拖曳的自包含暫定課表。"""
    fixed_keys = {
        (teacher, subject, klass, day, period)
        for teacher, subject, klass, day, period in fixed_assignments
    }
    binding_groups = {
        group.group_id: (binding.set_id, group)
        for binding in binding_sets
        for group in binding.groups
    }
    record_bindings = {}
    for group_id, (day, period, assigned) in binding_schedule.items():
        set_id, group = binding_groups[group_id]
        metadata = (set_id, group_id, group.mode)
        for klass, subject in assigned:
            record_bindings[(klass, subject, day, period, "R")] = metadata
        if group.special:
            teacher, subject, klass, role = group.special
            record_bindings[(teacher, subject, klass, day, period, role)] = metadata
    ordered = sorted(
        result,
        key=lambda row: (
            row[5] != "R",
            row[2],
            DAYS.index(row[3]),
            row[4],
            row[1],
            row[0],
        ),
    )
    records = []
    for index, (teacher, subject, klass, day, period, role) in enumerate(ordered, start=1):
        binding_metadata = record_bindings.get(
            (klass, subject, day, period, role) if role == "R"
            else (teacher, subject, klass, day, period, role)
        )
        locked_reason = ""
        if role == "R" and (teacher, subject, klass, day, period) in fixed_keys:
            locked_reason = "固定課"
        elif role == "R" and binding_metadata:
            locked_reason = "綁課"
        elif role == "S":
            locked_reason = "特生課"
        records.append({
            "id": f"L{index:04d}",
            "teacher": teacher,
            "subject": subject,
            "klass": klass,
            "day": day,
            "period": period,
            "role": role,
            "locked_reason": locked_reason,
            "binding_set_id": binding_metadata[0] if binding_metadata else "",
            "binding_group_id": binding_metadata[1] if binding_metadata else "",
            "binding_mode": binding_metadata[2] if binding_metadata else "",
        })

    serial_binding = {
        group_id: {
            "set_id": binding_groups[group_id][0],
            "mode": binding_groups[group_id][1].mode,
            "day": day,
            "period": period,
            "assigned": [[klass, subject] for klass, subject in assigned],
        }
        for group_id, (day, period, assigned) in binding_schedule.items()
    }
    payload = {
        "format": "junior-high-timetable-html-v1",
        "school_info": school_info,
        "days": DAYS,
        "periods": PERIODS,
        "period_times": {str(period): list(PERIOD_TIMES[period]) for period in PERIODS},
        "teacher_unavailable": {
            teacher: [[day, period] for day, period in sorted(slots, key=lambda slot: (DAYS.index(slot[0]), slot[1]))]
            for teacher, slots in teacher_unavailable.items()
        },
        "room_by_subject": ROOM_BY_SUBJECT,
        "physical_education_subject": PHYSICAL_EDUCATION_SUBJECT,
        "fixed_assignments": [list(assignment) for assignment in fixed_assignments],
        "binding_schedule": serial_binding,
        "result": records,
    }
    json_text = json.dumps(payload, ensure_ascii=False).replace("</", "<\\/")
    title = html.escape(f"{school_info['school_name']}－暫定課表微調", quote=True)
    document = ADJUSTMENT_HTML_TEMPLATE.replace("__TITLE__", title).replace("__DATA__", json_text)
    try:
        Path(filename).write_text(document, encoding="utf-8")
    except PermissionError:
        print(f"\n無法寫入 {filename}：檔案可能正在瀏覽器或編輯器中開啟，請關閉後再試。")
        return None
    return filename
